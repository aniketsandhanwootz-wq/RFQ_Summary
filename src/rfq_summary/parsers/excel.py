from __future__ import annotations

from io import BytesIO
from typing import Any, Dict, List
import re

from openpyxl import load_workbook  # type: ignore

from ..config import Settings
from ..schema import AttachmentFinding
from .image import analyze_image_bytes


def _cell_to_str(v) -> str:
    if v is None:
        return ""
    if isinstance(v, str):
        return v.strip()
    return str(v).strip()


def _extract_sheet_matrix(ws, max_rows: int, max_cols: int) -> List[List[str]]:
    rows: List[List[str]] = []
    max_r = min(ws.max_row or 0, max_rows)
    max_c = min(ws.max_column or 0, max_cols)
    for r in range(1, max_r + 1):
        row: List[str] = []
        for c in range(1, max_c + 1):
            row.append(_cell_to_str(ws.cell(r, c).value))
        rows.append(row)
    return rows


def _detect_table_regions(rows: List[List[str]], max_tables: int) -> List[Dict[str, Any]]:
    """
    Heuristic:
    - header row: >=2 non-empty and looks like labels (short strings)
    - table continues until empty separator row
    """
    tables = []
    r = 0
    n = len(rows)

    def looks_like_header(row: List[str]) -> bool:
        non_empty = [c for c in row if c]
        if len(non_empty) < 2:
            return False
        short = sum(1 for c in non_empty if len(c) <= 30)
        return short >= max(2, int(0.6 * len(non_empty)))

    def trim_right(rr: List[str]) -> List[str]:
        last = -1
        for i, c in enumerate(rr):
            if c:
                last = i
        return rr[: last + 1] if last >= 0 else []

    while r < n and len(tables) < max_tables:
        row = rows[r]
        if looks_like_header(row):
            header = trim_right(row)
            body = []
            r2 = r + 1
            while r2 < n:
                row2 = rows[r2]
                ne2 = sum(1 for c in row2 if c)
                if ne2 == 0:
                    break
                body.append(trim_right(row2))
                r2 += 1

            if header and body:
                tables.append(
                    {
                        "start_row": r + 1,
                        "end_row": r2,
                        "header": header[:40],
                        "rows_sample": [b[:40] for b in body[:200]],  # bigger sample for LLM
                        "row_count_sampled": min(len(body), 200),
                    }
                )
            r = r2 + 1
        else:
            r += 1

    return tables


def _find_formula_links(ws) -> List[Dict[str, str]]:
    links = []
    max_r = min(ws.max_row or 0, 300)
    max_c = min(ws.max_column or 0, 50)

    pat = re.compile(r"([A-Za-z0-9_ ]+)!([A-Z]{1,3}\d+)")
    for r in range(1, max_r + 1):
        for c in range(1, max_c + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.startswith("="):
                for m in pat.finditer(v):
                    links.append({"from_cell": ws.cell(r, c).coordinate, "to": f"{m.group(1)}!{m.group(2)}"})
                    if len(links) >= 120:
                        return links
    return links


def _rows_to_tsv(rows: List[List[str]], max_chars: int) -> str:
    """
    Convert sheet rows to a TSV-like text, keeping only non-empty rows.
    Bounded by max_chars to avoid context explosion.
    """
    out_lines: List[str] = []
    used = 0

    for rr in rows:
        if not any(c for c in rr):
            continue
        line = "\t".join(rr).rstrip()
        if not line:
            continue
        # hard bound
        if used + len(line) + 1 > max_chars:
            out_lines.append("...[TRUNCATED: sheet text exceeded budget]...")
            break
        out_lines.append(line)
        used += len(line) + 1

    return "\n".join(out_lines).strip()


def analyze_excel_bytes(settings: Settings, url: str, data: bytes) -> AttachmentFinding:
    # load values + formulas
    wb_vals = load_workbook(filename=BytesIO(data), data_only=True)
    wb_form = load_workbook(filename=BytesIO(data), data_only=False)

    sheet_summaries: List[Dict[str, Any]] = []
    embedded_images: List[Dict[str, Any]] = []
    relations: List[Dict[str, Any]] = []

    # Global text budget per Excel file (internal constant; no new env var)
    # Enough to include "entire excel" in practice for BOMs, while protecting context window.
    MAX_EXCEL_TEXT_CHARS = 180_000

    extracted_blocks: List[str] = []
    remaining = MAX_EXCEL_TEXT_CHARS

    for ws in wb_vals.worksheets:
        matrix = _extract_sheet_matrix(ws, settings.max_excel_rows, settings.max_excel_cols)
        tables = _detect_table_regions(matrix, settings.max_excel_tables_per_sheet)

        sheet_summaries.append(
            {
                "sheet": ws.title,
                "tables_detected": len(tables),
                "tables": tables,
            }
        )

        # --- build extracted text for this sheet ---
        if remaining > 0:
            extracted_blocks.append(f"## EXCEL SHEET: {ws.title}")
            remaining -= len(extracted_blocks[-1]) + 1

            # 1) table regions (best signal)
            if tables and remaining > 0:
                for ti, t in enumerate(tables, start=1):
                    if remaining <= 0:
                        break
                    hdr = t.get("header") or []
                    rows_sample = t.get("rows_sample") or []
                    block_lines = [f"[TABLE {ti}] start_row={t.get('start_row')} end_row={t.get('end_row')}"]
                    block_lines.append("\t".join([str(x) for x in hdr]))
                    for rr in rows_sample:
                        block_lines.append("\t".join([str(x) for x in rr]))
                    block = "\n".join(block_lines).strip()
                    if len(block) + 2 > remaining:
                        extracted_blocks.append("...[TRUNCATED: table text exceeded budget]...")
                        remaining = 0
                        break
                    extracted_blocks.append(block)
                    remaining -= len(block) + 2

            # 2) full sheet grid (non-empty rows) — ONLY if table detection is weak
            # We do this to avoid duplicating content and blowing up context.
            if remaining > 0:
                table_row_samples = 0
                if tables:
                    for tt in tables:
                        rs = tt.get("rows_sample") or []
                        table_row_samples += len(rs)

                # If we already captured enough table rows, skip grid dump.
                # Otherwise, include grid as a safety net.
                if table_row_samples < 20:
                    grid_text = _rows_to_tsv(matrix, max_chars=min(remaining, 70_000))
                    if grid_text:
                        block = "[SHEET_GRID_TSV]\n" + grid_text
                        if len(block) + 2 > remaining:
                            extracted_blocks.append("...[TRUNCATED: grid text exceeded budget]...")
                            remaining = 0
                        else:
                            extracted_blocks.append(block)
                            remaining -= len(block) + 2

        # embedded images (if any)
        imgs = getattr(ws, "_images", None) or []
        for idx, img in enumerate(imgs[:10]):
            try:
                img_bytes = img._data()  # type: ignore[attr-defined]
                finding = analyze_image_bytes(settings, f"{url}#sheet={ws.title}&img={idx+1}", img_bytes)
                embedded_images.append(
                    {
                        "sheet": ws.title,
                        "image_index": idx + 1,
                        "summary": finding.summary,
                    }
                )
            except Exception:
                embedded_images.append(
                    {
                        "sheet": ws.title,
                        "image_index": idx + 1,
                        "summary": "Embedded image detected but could not extract bytes.",
                    }
                )

    # formula relations
    for ws in wb_form.worksheets:
        links = _find_formula_links(ws)
        if links:
            relations.append({"sheet": ws.title, "formula_links": links[:120]})

    # compact summary text
    lines = [f"Excel analyzed. Sheets: {len(sheet_summaries)}."]
    for sh in sheet_summaries[:10]:
        lines.append(f"- Sheet '{sh['sheet']}': tables_detected={sh['tables_detected']}")
    if relations:
        lines.append(f"Cross-sheet formula links detected in {len(relations)} sheet(s).")
    if embedded_images:
        lines.append(f"Embedded images analyzed: {len(embedded_images)} (capped).")
    if extracted_blocks:
        lines.append(f"Extracted sheet text included for LLM (bounded to {MAX_EXCEL_TEXT_CHARS} chars).")

    extracted_text = "\n\n".join(extracted_blocks).strip()

    return AttachmentFinding(
        url=url,
        kind="excel",
        summary="\n".join(lines).strip(),
        data={
            "sheets": sheet_summaries,
            "relations": relations,
            "embedded_images": embedded_images,
            "extracted_text": extracted_text,  # <-- IMPORTANT: used by task.py
        },
    )