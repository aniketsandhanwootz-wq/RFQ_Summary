from __future__ import annotations

from typing import Dict, Any, List, Tuple
from io import BytesIO
import re

from openpyxl import load_workbook  # type: ignore

from ..config import Settings
from ..schema import AttachmentFinding
from .image import analyze_image_bytes


def _cell_to_str(v) -> str:
    if v is None:
        return ""
    if isinstance(v, str):
        return v.strip()
    return str(v).strip()


def _detect_table_regions(rows: List[List[str]]) -> List[Dict[str, Any]]:
    """
    Lightweight heuristic:
    - find first row with >= 2 non-empty cells => header
    - subsequent rows until a mostly-empty row => table body
    """
    tables = []
    r = 0
    n = len(rows)
    while r < n:
        row = rows[r]
        non_empty = [c for c in row if c]
        if len(non_empty) >= 2:
            header = row
            body = []
            r2 = r + 1
            while r2 < n:
                row2 = rows[r2]
                ne2 = sum(1 for c in row2 if c)
                # stop on mostly-empty separators
                if ne2 == 0:
                    break
                body.append(row2)
                r2 += 1

            # trim empty trailing cols
            def trim_right(rr):
                last = -1
                for i, c in enumerate(rr):
                    if c:
                        last = i
                return rr[: last + 1] if last >= 0 else []

            header_t = trim_right(header)
            body_t = [trim_right(b) for b in body][:50]  # cap rows to keep prompt compact

            # Accept only if header has some signal
            if len([c for c in header_t if c]) >= 2 and len(body_t) >= 1:
                tables.append(
                    {
                        "header": header_t[:30],
                        "rows_sample": [b[:30] for b in body_t],
                        "row_count_sampled": len(body_t),
                        "start_row": r + 1,
                        "end_row": r2,
                    }
                )
            r = r2 + 1
        else:
            r += 1
    return tables


def _extract_sheet_matrix(ws, max_rows: int = 200, max_cols: int = 30) -> List[List[str]]:
    rows: List[List[str]] = []
    max_r = min(ws.max_row or 0, max_rows)
    max_c = min(ws.max_column or 0, max_cols)
    for r in range(1, max_r + 1):
        row = []
        for c in range(1, max_c + 1):
            row.append(_cell_to_str(ws.cell(r, c).value))
        # keep even if empty; table detector handles separators
        rows.append(row)
    return rows


def analyze_excel_bytes(settings: Settings, url: str, data: bytes) -> AttachmentFinding:
    """
    Extract:
    - per-sheet matrix (sample)
    - table-like regions
    - embedded images (captions/OCR) with sheet anchoring
    """
    wb = load_workbook(filename=BytesIO(data), data_only=True)
    sheet_summaries: List[Dict[str, Any]] = []
    embedded_images: List[Dict[str, Any]] = []

    for ws in wb.worksheets:
        matrix = _extract_sheet_matrix(ws)
        tables = _detect_table_regions(matrix)

        sheet_summaries.append(
            {
                "sheet": ws.title,
                "tables_detected": len(tables),
                "tables": tables[:3],  # keep compact
            }
        )

        # Embedded images (openpyxl keeps them in ws._images)
        imgs = getattr(ws, "_images", None) or []
        for idx, img in enumerate(imgs[:10]):  # cap per sheet
            try:
                # openpyxl Image has ._data() in many versions
                img_bytes = img._data()  # type: ignore[attr-defined]
                finding = analyze_image_bytes(settings, f"{url}#sheet={ws.title}&img={idx+1}", img_bytes)
                embedded_images.append(
                    {
                        "sheet": ws.title,
                        "image_index": idx + 1,
                        "summary": finding.summary,
                    }
                )
            except Exception:
                embedded_images.append(
                    {
                        "sheet": ws.title,
                        "image_index": idx + 1,
                        "summary": "Embedded image detected but could not extract bytes for analysis.",
                    }
                )

    # Human-friendly summary
    lines = [f"Excel analyzed. Sheets: {len(sheet_summaries)}."]
    for sh in sheet_summaries[:6]:
        lines.append(f"- Sheet '{sh['sheet']}': tables_detected={sh['tables_detected']}")
    if embedded_images:
        lines.append(f"Embedded images analyzed: {len(embedded_images)} (capped).")

    return AttachmentFinding(
        url=url,
        kind="excel",
        summary="\n".join(lines).strip(),
        data={
            "sheets": sheet_summaries,
            "embedded_images": embedded_images,
        },
    )